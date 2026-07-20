import re
import io
import csv
import math
import difflib
from datetime import datetime
from collections import defaultdict

import openpyxl


def is_birthday_leave(desc):
    if not desc:
        return False
    s = str(desc).lower()
    s = re.sub(r'[^a-z0-9\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    if not s:
        return False
    compact = s.replace(' ', '')
    if 'birthday' in compact:
        return True
    tokens = s.split()
    for i, tok in enumerate(tokens):
        if tok == 'bday':
            return True
        if tok == 'b' and i + 1 < len(tokens) and tokens[i + 1] in ('day', 'days'):
            return True
    for n in (1, 2):
        for i in range(len(tokens) - n + 1):
            phrase = ''.join(tokens[i:i + n])
            if 6 <= len(phrase) <= 11:
                if difflib.SequenceMatcher(None, phrase, 'birthday').ratio() >= 0.8:
                    return True
    return False


def strip_paren(n):
    n = re.sub(r'\(.*?\)', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def name_key(n):
    return strip_paren(n).lower()


def first_last_key(n):
    parts = strip_paren(n).lower().split()
    if len(parts) < 2:
        return parts[0] if parts else ''
    return parts[0] + ' ' + parts[-1]


def fmt_date(d):
    return d.strftime('%d-%b-%Y')


def fmt_date_list(dates):
    if not dates:
        return 'None'
    return ', '.join(fmt_date(d) for d in sorted(dates))


def _sniff_delimiter(first_line):
    """WTC dumps have been seen both comma-delimited and semicolon-delimited
    (semicolon dumps also quote every field). Pick whichever delimiter is
    more common on the header line."""
    return ';' if first_line.count(';') > first_line.count(',') else ','


def load_wtc(file_obj, hours_per_day=8.0):
    if hasattr(file_obj, 'read'):
        raw = file_obj.read()
        if isinstance(raw, bytes):
            raw = raw.decode('utf-8', errors='replace')
    else:
        with open(file_obj, encoding='utf-8', errors='replace') as fh:
            raw = fh.read()

    first_line = raw.split('\n', 1)[0]
    delimiter = _sniff_delimiter(first_line)
    text_stream = io.StringIO(raw)

    wtc_bday_entries = defaultdict(int)
    wtc_display_name = {}
    wtc_bday_flag = defaultdict(bool)
    wtc_date_days = defaultdict(lambda: defaultdict(float))

    reader = csv.DictReader(text_stream, delimiter=delimiter)
    for row in reader:
        if row.get('db') != 'occ_leaves':
            continue
        uname = (row.get('user_name') or '').strip()
        if not uname:
            continue
        key = name_key(uname)
        wtc_display_name.setdefault(key, strip_paren(uname))
        try:
            hrs = float(row['duration_hrs']) if row.get('duration_hrs') not in (None, '') else 0.0
        except ValueError:
            hrs = 0.0
        days = hrs / hours_per_day

        raw_date = (row.get('date') or '').split(' ')[0].strip()
        parsed_date = None
        if raw_date:
            try:
                parsed_date = datetime.strptime(raw_date, '%m/%d/%Y').date()
            except ValueError:
                parsed_date = None

        desc = row.get('description', '') or ''
        if is_birthday_leave(desc):
            wtc_bday_entries[key] += 1
            wtc_bday_flag[key] = True
        else:
            if parsed_date is not None:
                wtc_date_days[key][parsed_date] += days

    return wtc_bday_entries, wtc_bday_flag, wtc_display_name, wtc_date_days


def load_workday(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [c.value for c in ws[1]]
    worker_col = header.index('Worker') + 1
    approved_col = header.index('Approved') + 1
    date_col = header.index('Time Off Date') + 1
    type_col = header.index('Type') + 1 if 'Type' in header else None
    entry_col = header.index('Time Off Entry') + 1 if 'Time Off Entry' in header else None
    manager_col = header.index('Manager') + 1 if 'Manager' in header else None

    groups = {}
    wd_display_name = {}
    wd_manager = {}

    for r in range(2, ws.max_row + 1):
        worker = ws.cell(r, worker_col).value
        if not worker:
            continue
        worker = str(worker).strip()
        key = name_key(worker)
        wd_display_name.setdefault(key, strip_paren(worker))

        if manager_col:
            manager_val = ws.cell(r, manager_col).value
            if manager_val and str(manager_val).strip():
                wd_manager.setdefault(key, strip_paren(str(manager_val).strip()))

        entry_text = ws.cell(r, entry_col).value if entry_col else None
        leave_type = ws.cell(r, type_col).value if type_col else None
        group_key = (key, entry_text, leave_type)

        raw_date = ws.cell(r, date_col).value
        parsed_date = None
        if raw_date:
            try:
                parsed_date = datetime.strptime(str(raw_date).strip(), '%d/%m/%Y').date()
            except ValueError:
                parsed_date = None

        val = ws.cell(r, approved_col).value
        try:
            num_val = float(val) if val not in (None, '') else None
        except (ValueError, TypeError):
            num_val = None

        g = groups.setdefault(group_key, {'display': worker, 'date': parsed_date, 'vals': []})
        g['vals'].append(num_val)
        if g['date'] is None:
            g['date'] = parsed_date

    wd_date_days = defaultdict(lambda: defaultdict(float))

    for (key, entry_text, leave_type), g in groups.items():
        # Approved Leave Rule: if Approved is blank/None/0 for every duplicate
        # row in this group, the leave record is not a valid approved leave -
        # exclude it entirely from the comparison (don't estimate from text).
        valid_vals = [v for v in g['vals'] if v not in (None, 0, 0.0)]
        if not valid_vals:
            continue
        days = valid_vals[0]

        if g['date'] is not None:
            wd_date_days[key][g['date']] += days

    return wd_display_name, wd_date_days, wd_manager


def build_report(wtc_file, workday_file, hours_per_day=8.0):
    (wtc_bday_entries, wtc_bday_flag,
     wtc_display_name, wtc_date_days) = load_wtc(wtc_file, hours_per_day)
    (wd_display_name, wd_date_days, wd_manager) = load_workday(workday_file)

    all_wtc_keys = set(wtc_display_name.keys())
    all_wd_keys = set(wd_display_name.keys())

    used_wd, used_wtc = set(), set()
    matched_pairs = []

    for k in sorted(all_wtc_keys & all_wd_keys):
        matched_pairs.append((k, k))
        used_wtc.add(k)
        used_wd.add(k)

    remaining_wtc = all_wtc_keys - used_wtc
    remaining_wd = all_wd_keys - used_wd
    wd_fl_index = defaultdict(list)
    for k in remaining_wd:
        wd_fl_index[first_last_key(wd_display_name[k])].append(k)

    fuzzy_note = {}
    for k in sorted(remaining_wtc):
        fl = first_last_key(wtc_display_name[k])
        candidates = wd_fl_index.get(fl, [])
        if len(candidates) == 1:
            wdk = candidates[0]
            matched_pairs.append((k, wdk))
            used_wtc.add(k)
            used_wd.add(wdk)
            fuzzy_note[(k, wdk)] = 'Name variant matched (middle name/suffix differs)'

    remaining_wd = all_wd_keys - used_wd
    remaining_wtc = all_wtc_keys - used_wtc
    for k in sorted(remaining_wtc):
        matched_pairs.append((k, None))
    for k in sorted(remaining_wd):
        matched_pairs.append((None, k))

    def dates_with_leave(date_days_map, key):
        return {d for d, amt in date_days_map.get(key, {}).items() if amt > 1e-9}

    rows_out = []
    for wtc_key, wd_key in matched_pairs:
        if wtc_key and wd_key:
            display = wtc_display_name[wtc_key]
            lead_name = wd_manager.get(wd_key, 'N/A')
            bday_flag = wtc_bday_flag.get(wtc_key, False)
            bday_entries = wtc_bday_entries.get(wtc_key, 0)

            # Leave Count Rule: a "leave entry" is a calendar date with a net
            # positive leave amount, after netting all same-date records
            # (including cancellations/corrections) together. A date that is
            # booked and then fully cancelled (net <= 0) is not an entry; a
            # date that is cancelled and rebooked (net > 0) still counts as
            # exactly one entry - never double-counted for the same date.
            wtc_dates = dates_with_leave(wtc_date_days, wtc_key)
            wd_dates = dates_with_leave(wd_date_days, wd_key)
            wtc_entries = len(wtc_dates)
            wd_entries = len(wd_dates)
            entries_deviation = wtc_entries - wd_entries

            missing_in_workday = wtc_dates - wd_dates
            missing_in_wtc = wd_dates - wtc_dates

            remark_bits = []
            if (wtc_key, wd_key) in fuzzy_note:
                remark_bits.append(fuzzy_note[(wtc_key, wd_key)])
            if entries_deviation == 0:
                base_remark = 'No deviation in leave entries'
            elif entries_deviation > 0:
                base_remark = f'{entries_deviation} leave entry(ies) missing in Workday'
            else:
                base_remark = f'{abs(entries_deviation)} leave entry(ies) extra in Workday (not in WTC)'
            if bday_flag:
                base_remark += f' ({bday_entries} Bday leave entry(ies) excluded)'
            remark_bits.append(base_remark)

            rows_out.append({
                'User Name': display,
                'Lead Name': lead_name,
                'WTC Leave Entries': wtc_entries,
                'Workday Leave Entries': wd_entries,
                'Entries Deviation': entries_deviation,
                'Birthday Leave (Yes/No)': 'Yes' if bday_flag else 'No',
                'Remarks': '; '.join(remark_bits),
                'Missing Dates in Workday': fmt_date_list(missing_in_workday),
                'Missing Dates in WTC': fmt_date_list(missing_in_wtc),
            })
        elif wtc_key and not wd_key:
            display = wtc_display_name[wtc_key]
            bday_flag = wtc_bday_flag.get(wtc_key, False)
            bday_entries = wtc_bday_entries.get(wtc_key, 0)
            wtc_dates = dates_with_leave(wtc_date_days, wtc_key)
            wtc_entries = len(wtc_dates)
            remark = 'No matching Workday record found for this employee'
            if bday_flag:
                remark += f' ({bday_entries} Bday leave entry(ies) excluded)'
            rows_out.append({
                'User Name': display,
                'Lead Name': 'N/A',
                'WTC Leave Entries': wtc_entries,
                'Workday Leave Entries': 'N/A',
                'Entries Deviation': 'N/A',
                'Birthday Leave (Yes/No)': 'Yes' if bday_flag else 'No',
                'Remarks': remark,
                'Missing Dates in Workday': fmt_date_list(wtc_dates),
                'Missing Dates in WTC': 'N/A - no Workday record',
            })
        else:
            display = wd_display_name[wd_key]
            lead_name = wd_manager.get(wd_key, 'N/A')
            wd_dates = dates_with_leave(wd_date_days, wd_key)
            wd_entries = len(wd_dates)
            rows_out.append({
                'User Name': display,
                'Lead Name': lead_name,
                'WTC Leave Entries': 'N/A',
                'Workday Leave Entries': wd_entries,
                'Entries Deviation': 'N/A',
                'Birthday Leave (Yes/No)': 'No',
                'Remarks': 'No matching WTC record found for this employee',
                'Missing Dates in Workday': 'N/A - no WTC record',
                'Missing Dates in WTC': fmt_date_list(wd_dates),
            })

    def sort_key(r):
        d = r['Entries Deviation']
        if d == 'N/A':
            return (1, 0)
        return (0, -abs(d))

    rows_out.sort(key=sort_key)
    return rows_out


def _needed_lines(text, col_width):
    """Rough estimate of how many wrapped lines a cell's text will take at a
    given column width, so we can grow the row height and avoid Excel
    visually clipping long wrapped text (e.g. long Missing Dates lists)."""
    if not text:
        return 1
    chars_per_line = max(int(col_width), 10)
    return max(1, math.ceil(len(str(text)) / chars_per_line))


def save_report_xlsx(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Leave Deviation"

    FONT_NAME = "Arial"
    headers = ["User Name", "Lead Name", "WTC Leave Entries", "Workday Leave Entries",
               "Entries Deviation", "Birthday Leave (Yes/No)", "Remarks",
               "Missing Dates in Workday", "Missing Dates in WTC"]
    ws.append(headers)

    header_fill = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    red_fill = PatternFill("solid", start_color="FF7C80", end_color="FF7C80")
    green_fill = PatternFill("solid", start_color="C6E0B4", end_color="C6E0B4")
    ncols = len(headers)

    widths = [24, 22, 16, 18, 16, 18, 42, 32, 32]

    r = 2
    for row in rows:
        wtc_e, wd_e = row["WTC Leave Entries"], row["Workday Leave Entries"]
        entries_dev = row["Entries Deviation"]
        bday_flag = row["Birthday Leave (Yes/No)"]
        remark = row["Remarks"]
        missing_wd, missing_wtc = row["Missing Dates in Workday"], row["Missing Dates in WTC"]

        ws.cell(r, 1, row["User Name"])
        ws.cell(r, 2, row["Lead Name"])
        ws.cell(r, 3, wtc_e)
        ws.cell(r, 4, wd_e)
        if isinstance(wtc_e, (int, float)) and isinstance(wd_e, (int, float)):
            ws.cell(r, 5, f"=C{r}-D{r}")
        else:
            ws.cell(r, 5, "N/A")
        ws.cell(r, 6, bday_flag)
        ws.cell(r, 7, remark)
        ws.cell(r, 8, missing_wd)
        ws.cell(r, 9, missing_wtc)

        for c in range(1, ncols + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name=FONT_NAME, size=10.5)
            cell.border = border
            if c in (3, 4, 5, 6):
                cell.alignment = Alignment(horizontal="center")
            elif c in (7, 8, 9):
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Grow the row height so wrapped Remarks/Missing-Dates text is never
        # visually clipped by Excel's default row height.
        lines = max(
            _needed_lines(remark, widths[6]),
            _needed_lines(missing_wd, widths[7]),
            _needed_lines(missing_wtc, widths[8]),
            1,
        )
        ws.row_dimensions[r].height = max(15, lines * 14)

        is_na = entries_dev == "N/A"
        has_real_dev = (not is_na) and isinstance(entries_dev, (int, float)) and entries_dev != 0
        if has_real_dev:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = red_fill
        elif bday_flag == "Yes" and not is_na:
            for c in range(1, ncols + 1):
                ws.cell(r, c).fill = green_fill
        r += 1

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{r - 1}"
    wb.save(out_path)
    return out_path
