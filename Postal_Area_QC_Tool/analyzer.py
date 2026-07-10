# =============================================================================
# Copyright (c) 2025 ADP Team, TomTom. All rights reserved.
# Contact: prabhakar.chaudhari@tomtom.com | sachin.shete@tomtom.com
# =============================================================================
"""
Postal Area Quality Check -- Core Analysis Logic  (optimised build)

Step 1  Spike Detection    (spike 1.fmw)            -> POINT  "after_topo_tool"
Step 2  Sliver Detection   (Single_Layer_Sliver_Polygon_Finder 1.fmw) -> POLYGON
Step 3  Geometry Failure Check -> POLYGON "Geometry_Failed_Less/Greater_Than_200m"
Step 4  Multipart Analysis (ArcGIS QC)              -> POLYGON "Multipart_Validation"
"""

from __future__ import annotations

import math
import os
import datetime
from dataclasses import dataclass
from typing import Optional, Tuple, List

import numpy as np
import geopandas as gpd
import pandas as pd
from shapely.geometry import Point, MultiPolygon, MultiLineString, LineString
from shapely.ops import unary_union, polygonize
from shapely.strtree import STRtree


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class QCResult:
    spike:            gpd.GeoDataFrame   # POINT   -- Spike_Output
    sliver:           gpd.GeoDataFrame   # POLYGON -- Sliver_Suspicious_Area_Output
    geom_lt1:         gpd.GeoDataFrame   # POLYGON -- Geometry_Failed_Less_Than_1m  (fixed)
    geom_lt_buf:      gpd.GeoDataFrame   # POLYGON -- Geometry_Failed_Less_Than_<threshold>m
    geom_gt_buf:      gpd.GeoDataFrame   # POLYGON -- Geometry_Failed_Greater_Than_<threshold>m
    multipart:        gpd.GeoDataFrame   # POLYGON -- Multipart_Validation (individual
                                         #            unmatched parts, not whole features)
    poly_quality:     gpd.GeoDataFrame   # POLYGON -- Polygon_Quality_Check: sym-diff
                                         #            parts with deviation > 5 m (hardcoded)
    summary:          dict
    timestamp:        str


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def list_layers(path: str) -> List[str]:
    import fiona
    if path.lower().endswith(".shp"):
        return [os.path.splitext(os.path.basename(path))[0]]
    return fiona.listlayers(path)


def read_layer(path: str, layer: Optional[str] = None) -> gpd.GeoDataFrame:
    import fiona
    if path.lower().endswith(".shp"):
        gdf = gpd.read_file(path)
    else:
        layers = fiona.listlayers(path)
        if not layers:
            raise ValueError("No layers found in " + path)
        chosen = layer if (layer and layer in layers) else layers[0]
        gdf = gpd.read_file(path, layer=chosen)
    if gdf.crs is None:
        gdf = gdf.set_crs("EPSG:4326")
    elif gdf.crs.to_epsg() != 4326:
        gdf = gdf.to_crs("EPSG:4326")
    return gdf


# Standard attribute columns to carry/compute in every output layer
_STD_COLS = ["uuid", "CP5", "ddctType", "X", "Y", "SHAPE_Area", "SHAPE_Length"]


def _reorder_cols(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Put standard columns first (if present), then the rest, then geometry."""
    existing_std = [c for c in _STD_COLS if c in gdf.columns]
    other = [c for c in gdf.columns if c not in _STD_COLS and c != "geometry"]
    new_order = existing_std + other + ["geometry"]
    return gdf[[c for c in new_order if c in gdf.columns]]


def _best_utm(gdf: gpd.GeoDataFrame) -> int:
    """Return best UTM EPSG for a GeoDataFrame using total_bounds (no topology ops)."""
    minx, miny, maxx, maxy = gdf.total_bounds
    cx = (minx + maxx) / 2
    cy = (miny + maxy) / 2
    zone = int((cx + 180) / 6) + 1
    return (32600 + zone) if cy >= 0 else (32700 + zone)


def _make_valid_gdf(gdf: gpd.GeoDataFrame) -> gpd.GeoDataFrame:
    """Return a copy of gdf with all invalid geometries repaired.

    Uses buffer(0) first (fast), then shapely.make_valid as fallback.
    Null/empty geometries are dropped.
    """
    gdf = gdf[gdf.geometry.notna() & ~gdf.geometry.is_empty].copy()
    invalid_mask = ~gdf.geometry.is_valid
    if not invalid_mask.any():
        return gdf
    try:
        from shapely.validation import make_valid
        gdf.loc[invalid_mask, "geometry"] = gdf.loc[invalid_mask, "geometry"].apply(make_valid)
    except ImportError:
        gdf.loc[invalid_mask, "geometry"] = gdf.loc[invalid_mask, "geometry"].buffer(0)
    return gdf


# ---------------------------------------------------------------------------
# Step 1  Spike Detection  (numpy-vectorised per ring)
# Matches FME: SpikeRemoverFactory SPIKE_ANGLE=15, output = after_topo_tool
# SHAPE_Area = parent polygon area in WGS84 geographic units (degrees sq)
# ---------------------------------------------------------------------------

def _ring_spike_angles(coords: np.ndarray, threshold: float):
    """Return (spike_indices, all_angles) for a ring coordinate array.

    Matches FME SpikeRemoverFactory SPIKE_ANGLE=15:
      - interior angle <= threshold is a spike  (FME: "maximum angle")
      - lone triangles (3 distinct vertices) are NOT spikes  — FME logs
        "SpikeRemover: Lone triangles are not considered spikes."
      - degenerate vertices (zero-length adjacent edge) are skipped to avoid
        false positives from near-duplicate ring vertices
    """
    n = len(coords)
    if n < 3:
        return np.empty(0, dtype=int), np.empty(0)
    # FME explicitly skips lone triangles (rings with exactly 3 distinct vertices)
    if n == 3:
        return np.empty(0, dtype=int), np.empty(0)
    p1 = coords[(np.arange(n) - 1) % n]
    p2 = coords
    p3 = coords[(np.arange(n) + 1) % n]
    ax = p1[:, 0] - p2[:, 0]
    ay = p1[:, 1] - p2[:, 1]
    bx = p3[:, 0] - p2[:, 0]
    by = p3[:, 1] - p2[:, 1]
    # Skip vertices where either adjacent edge has near-zero length (degenerate /
    # duplicate vertices).  A zero-length edge makes atan2(0,0) = 0 which would
    # always trigger as a false spike.
    MIN_EDGE = 1e-9
    len_a = np.sqrt(ax * ax + ay * ay)
    len_b = np.sqrt(bx * bx + by * by)
    valid = (len_a >= MIN_EDGE) & (len_b >= MIN_EDGE)
    cross  = np.abs(ax * by - ay * bx)
    dot    = ax * bx + ay * by
    angles = np.degrees(np.arctan2(cross, dot))
    # FME SPIKE_ANGLE=15 means "maximum interior angle" → flag if angle <= threshold
    spike_mask = valid & (angles <= threshold)
    return np.where(spike_mask)[0], angles


def detect_spikes(gdf: gpd.GeoDataFrame,
                  spike_angle: float = 15.0) -> gpd.GeoDataFrame:
    """Detect spike vertices on MDS polygons.

    Matches FME SpikeRemoverFactory SPIKE_ANGLE=15, single-pass on original
    ring coordinates.  REMOVE_SPIKES_ITERATIVELY=Yes in the FME workspace
    applies to cleaning the OUTPUT polygon geometry, not to spike detection;
    the REMOVED port (spike points) reflects a single pass on the original ring.
    """
    point_rows = []
    attr_cols  = [c for c in gdf.columns if c != "geometry"]
    # Vectorised extraction: avoids slow iterrows() per-row overhead.
    _geom_vals  = gdf.geometry.values
    _area_vals  = gdf.geometry.area.values
    _len_vals   = gdf.geometry.length.values
    _attr_arrs  = {c: gdf[c].values for c in attr_cols}
    for _i in range(len(gdf)):
        geom = _geom_vals[_i]
        if geom is None or geom.is_empty:
            continue
        polys = list(geom.geoms) if geom.geom_type == "MultiPolygon" else [geom]
        attrs = {c: _attr_arrs[c][_i] for c in attr_cols}
        # Vectorised area/length (already computed above)
        attrs["SHAPE_Area"]   = float(_area_vals[_i])
        attrs["SHAPE_Length"] = float(_len_vals[_i])
        for poly in polys:
            for ring in [poly.exterior] + list(poly.interiors):
                pts = np.array(ring.coords)
                if len(pts) > 1 and np.array_equal(pts[0], pts[-1]):
                    pts = pts[:-1]
                idxs, angles = _ring_spike_angles(pts, spike_angle)
                for i in idxs:
                    rec = dict(attrs)
                    rec["_spike_angle"] = round(float(angles[i]), 4)
                    rec["X"]            = float(pts[i, 0])
                    rec["Y"]            = float(pts[i, 1])
                    rec["geometry"]     = Point(pts[i])
                    point_rows.append(rec)
    if not point_rows:
        return gpd.GeoDataFrame(
            columns=attr_cols + ["SHAPE_Area", "SHAPE_Length",
                                 "_spike_angle", "X", "Y", "geometry"],
            geometry="geometry", crs=gdf.crs)
    out = gpd.GeoDataFrame(point_rows, geometry="geometry",
                           crs=gdf.crs).reset_index(drop=True)
    return _reorder_cols(out)


# ---------------------------------------------------------------------------
# Step 2  Sliver Detection
# Matches FME Single_Layer_Sliver_Polygon_Finder:
#   1. Node all MDS polygon boundaries via unary_union
#   2. Polygonize the noded line network → planar graph polygons
#   3. Flag built polygons NOT containing any original feature centroid
#      (= gap/sliver polygons between adjacent features)
#   4. Keep only those where round(area, 6) == 0  (< ~12 m² in WGS84)
# Output: POLYGON layer  "Sliver_Suspicious_Area_Output"
# ---------------------------------------------------------------------------

def detect_slivers(gdf: gpd.GeoDataFrame,
                   area_decimals: int = 6,
                   progress_callback=None) -> gpd.GeoDataFrame:
    valid = _make_valid_gdf(gdf).reset_index(drop=True)
    attr_cols = [c for c in valid.columns if c != "geometry"]

    empty_cols = attr_cols + ["X", "Y", "SHAPE_Area", "SHAPE_Length", "geometry"]

    if len(valid) == 0:
        return gpd.GeoDataFrame(columns=empty_cols, geometry="geometry", crs=gdf.crs)

    # --- Step 1: node all boundaries ---
    if progress_callback:
        progress_callback("Sliver: noding polygon boundaries…")
    boundaries = [geom.boundary for geom in valid.geometry]
    noded_lines = unary_union(boundaries)

    # --- Step 2: polygonize the noded line network ---
    if progress_callback:
        progress_callback("Sliver: polygonizing noded boundaries…")
    built_polys = list(polygonize(noded_lines))
    if not built_polys:
        return gpd.GeoDataFrame(columns=empty_cols, geometry="geometry", crs=gdf.crs)

    # --- Step 3: spatial index of original centroids (ownership test) ---
    centroids  = valid.geometry.centroid.values
    ctree      = STRtree(centroids)

    # --- Step 4: spatial index of MDS geometries (nearest-feature lookup) ---
    geom_tree  = STRtree(valid.geometry.values)

    # --- Step 5: identify gap polygons and attach nearest MDS attributes ---
    if progress_callback:
        progress_callback("Sliver: identifying gap polygons…")
    sliver_rows = []
    for poly in built_polys:
        if poly is None or poly.is_empty:
            continue
        # Ownership test: skip if any original centroid falls inside
        cand_idxs = ctree.query(poly)
        owned = any(poly.contains(centroids[j]) for j in cand_idxs)
        if owned:
            continue
        area = poly.area
        if round(area, area_decimals) != 0:
            continue

        # Find the nearest MDS feature to carry its attributes
        cen = poly.centroid
        near_idxs = geom_tree.nearest(cen)
        # nearest() may return int or array
        if hasattr(near_idxs, '__len__'):
            near_idx = int(near_idxs[0])
        else:
            near_idx = int(near_idxs)

        nearest_row = valid.iloc[near_idx]
        rec = {c: nearest_row[c] for c in attr_cols}
        rec["X"]            = round(cen.x, 9)
        rec["Y"]            = round(cen.y, 9)
        rec["SHAPE_Area"]   = area
        rec["SHAPE_Length"] = poly.length
        rec["geometry"]     = poly
        sliver_rows.append(rec)

    if not sliver_rows:
        return gpd.GeoDataFrame(columns=empty_cols, geometry="geometry", crs=gdf.crs)

    out = gpd.GeoDataFrame(sliver_rows, geometry="geometry",
                           crs=gdf.crs).reset_index(drop=True)
    return _reorder_cols(out)


# ---------------------------------------------------------------------------
# Step 3  Geometry Failure Check
#
# Logic:
#   1. Match each MDS polygon to its Source polygon by shared ID (uuid / CP5).
#   2. Compute symmetric_difference(source_geom, mds_geom).
#   3. Decompose the result into individual polygon parts.
#   4. Project each part to UTM; measure the minimum width of the minimum
#      rotated bounding rectangle (MRR).  This equals the perpendicular shift
#      between the two edges — correct for long thin boundary-shift slivers.
#      (Old bbox-max would report the LENGTH of the sliver, not the shift.)
#
# Two POLYGON outputs:
#   Geometry_Failed_Less_Than_Xm    — parts with deviation <= threshold
#   Geometry_Failed_Greater_Than_Xm — parts with deviation >  threshold
#
# Each row carries both mds_* and src_* attribute columns.
# ---------------------------------------------------------------------------

_GEOM_NOISE_FILTER_M        = 1.0    # metres — parts smaller than this are coordinate rounding artifacts
_GEOM_LINEAR_ASPECT_RATIO   = 4.0    # long/short > this (combined with width+length) → linear artefact
_GEOM_ROAD_WIDTH_MAX_M      = 30.0   # metres — slivers narrower than this are candidate road artefacts
_GEOM_SLIVER_MIN_LENGTH_M   = 150.0  # metres — road slivers must be at least this long to be filtered;
                                     #          short narrow features (<150 m) are kept as real errors


def _effective_threshold(buffer_dist_m: float) -> float:
    """Use the user-supplied buffer value directly as the split threshold.
    The 1 m noise filter is always fixed and independent of this.
    """
    return float(buffer_dist_m)


def _step3_layer_names(threshold_m: float):
    """Return (lt_name, gt_name) for the given effective threshold."""
    t = int(threshold_m) if threshold_m == int(threshold_m) else threshold_m
    return (
        f"Geometry_Failed_Less_Than_{t}m",
        f"Geometry_Failed_Greater_Than_{t}m",
    )


_DEDUP_IOU_THRESHOLD = 0.95   # parts with IoU >= this are considered shared-boundary duplicates


def _dedup_sym_diff_rows(rows: list) -> list:
    """Remove near-duplicate sym-diff parts caused by shared polygon boundaries.

    When two adjacent features share a boundary that changed, the symmetric
    difference of each feature contains the same geometric sliver — effectively
    counting the same shift twice (or more).  This function keeps only the first
    occurrence and drops any subsequent part whose IoU (intersection / union area)
    against an already-accepted part is >= _DEDUP_IOU_THRESHOLD.

    Uses a single STRtree pass: O(n * k) where k = average candidate neighbours.
    """
    if len(rows) <= 1:
        return rows

    geoms  = [r["geometry"] for r in rows]
    tree   = STRtree(geoms)
    dropped = set()

    for i, g in enumerate(geoms):
        if i in dropped:
            continue
        for j in tree.query(g):
            if j <= i or j in dropped:
                continue
            kg = geoms[j]
            if not g.intersects(kg):
                continue
            try:
                inter = g.intersection(kg).area
                union = g.area + kg.area - inter
                if union > 0 and inter / union >= _DEDUP_IOU_THRESHOLD:
                    dropped.add(j)
            except Exception:
                continue

    return [r for i, r in enumerate(rows) if i not in dropped]


def _geom_utm_epsg(geom) -> int:
    """Return best UTM EPSG for a single shapely geometry."""
    cen = geom.centroid
    zone = int((cen.x + 180) / 6) + 1
    return (32600 + zone) if cen.y >= 0 else (32700 + zone)


def buffer_comparison(
    source_gdf:       gpd.GeoDataFrame,
    mds_gdf:          gpd.GeoDataFrame,
    buffer_dist_m:    float = 200.0,
    progress_callback=None,
    src_id_col:       Optional[str] = None,
    mds_id_col:       Optional[str] = None,
) -> Tuple[gpd.GeoDataFrame, gpd.GeoDataFrame, gpd.GeoDataFrame]:
    """
    For each MDS/Source matched pair (by shared ID):
      1. Compute symmetric_difference(source_geom, mds_geom).
      2. Decompose into individual polygon parts.
      3. Project each part to UTM; measure bounding-box max dimension.
      4. Always split at 1 m (noise filter); then split at _effective_threshold(buffer_dist_m).

    src_id_col / mds_id_col: explicit column names to use for matching.
    If not provided, _find_id_col_pair() is called to auto-detect them
    (supports same-name columns AND different-name columns with overlapping values).

    Returns:
      lt1_gdf    — Geometry_Failed_Less_Than_1m        (deviation < 1 m — noise, fixed)
      lt_buf_gdf — Geometry_Failed_Less_Than_<t>m      (1 m <= deviation <= threshold)
      gt_buf_gdf — Geometry_Failed_Greater_Than_<t>m   (deviation > threshold)

    Each row carries both mds_* and src_* attribute columns.
    """
    from pyproj import Transformer
    from shapely.ops import transform as shp_transform

    src = source_gdf.to_crs(epsg=4326).copy()
    mds = mds_gdf.to_crs(epsg=4326).copy()

    # Resolve ID columns (same name or different name via value-overlap detection)
    _src_id, _mds_id = _find_id_col_pair(
        src, mds,
        src_id_col=src_id_col,
        mds_id_col=mds_id_col,
    )
    if _src_id is None or _mds_id is None:
        import warnings
        warnings.warn("Step 3 skipped: no common ID column found between Source and MDS.")
        empty = gpd.GeoDataFrame(columns=["geometry"], geometry="geometry", crs="EPSG:4326")
        return empty, empty.copy(), empty.copy()

    src_ids    = set(src[_src_id].dropna().astype(str))
    mds_ids    = set(mds[_mds_id].dropna().astype(str))
    common_ids = src_ids & mds_ids

    # Build index keyed by the common ID values (string)
    src_idx = src[src[_src_id].astype(str).isin(common_ids)].copy()
    src_idx["_id_key"] = src_idx[_src_id].astype(str)
    src_idx = src_idx.set_index("_id_key")
    src_idx = src_idx[~src_idx.index.duplicated(keep="first")]

    mds_idx = mds[mds[_mds_id].astype(str).isin(common_ids)].copy()
    mds_idx["_id_key"] = mds_idx[_mds_id].astype(str)
    mds_idx = mds_idx.set_index("_id_key")
    mds_idx = mds_idx[~mds_idx.index.duplicated(keep="first")]

    mds_attr_cols = [c for c in mds.columns if c not in ("geometry", _mds_id)]
    src_attr_cols = [c for c in src.columns if c not in ("geometry", _src_id)]

    src_geom_map  = dict(zip(src_idx.index.astype(str), src_idx["geometry"]))
    mds_geom_map  = dict(zip(mds_idx.index.astype(str), mds_idx["geometry"]))
    # Filter attr cols to only those actually present after indexing
    src_attr_cols = [c for c in src_attr_cols if c in src_idx.columns]
    mds_attr_cols = [c for c in mds_attr_cols if c in mds_idx.columns]
    src_attrs_map = src_idx[src_attr_cols].to_dict(orient="index") if src_attr_cols else {}
    mds_attrs_map = mds_idx[mds_attr_cols].to_dict(orient="index") if mds_attr_cols else {}

    # Cache UTM Transformer objects by EPSG zone
    _t_cache: dict = {}
    def _get_t(epsg: int):
        if epsg not in _t_cache:
            _t_cache[epsg] = Transformer.from_crs("EPSG:4326", f"EPSG:{epsg}",
                                                   always_xy=True)
        return _t_cache[epsg]

    eff_threshold = _effective_threshold(buffer_dist_m)
    lt1_rows      = []   # < 1 m        — noise (fixed)
    lt_buf_rows   = []   # 1 m-threshold — minor
    gt_buf_rows   = []   # > threshold   — major
    n = len(common_ids)

    for i, fid in enumerate(common_ids):
        if progress_callback and i % max(1, n // 20) == 0:
            progress_callback(i, n, f"Step 3: comparing feature {i:,}/{n:,}...")

        src_geom = src_geom_map.get(fid)
        mds_geom = mds_geom_map.get(fid)
        if src_geom is None or mds_geom is None:
            continue
        if src_geom.equals(mds_geom):
            continue   # identical — no unmatched geometry

        # Fix invalid geometries before computing symmetric difference.
        # TopologyException is raised by GEOS when input has self-intersections
        # or other topology errors (common in raw postal/admin boundary data).
        try:
            if not src_geom.is_valid:
                src_geom = src_geom.buffer(0)
            if not mds_geom.is_valid:
                mds_geom = mds_geom.buffer(0)
            sym_diff = src_geom.symmetric_difference(mds_geom)
        except Exception:
            # Last resort: make_valid (Shapely 1.8+ / GEOS 3.8+)
            try:
                from shapely.validation import make_valid
                src_geom = make_valid(src_geom)
                mds_geom = make_valid(mds_geom)
                sym_diff = src_geom.symmetric_difference(mds_geom)
            except Exception:
                continue   # skip this feature pair — cannot fix geometry
        if sym_diff is None or sym_diff.is_empty:
            continue

        # Decompose sym-diff into individual polygon parts
        if sym_diff.geom_type == "Polygon":
            parts = [sym_diff]
        elif sym_diff.geom_type in ("MultiPolygon", "GeometryCollection"):
            parts = [g for g in sym_diff.geoms
                     if g.geom_type == "Polygon" and not g.is_empty]
        else:
            parts = []

        mds_attrs = mds_attrs_map.get(fid, {})
        src_attrs = src_attrs_map.get(fid, {})
        utm_epsg  = _geom_utm_epsg(src_geom)

        for part in parts:
            max_dim_m    = 0.0
            aspect_ratio = 1.0
            try:
                part_utm = shp_transform(_get_t(utm_epsg).transform, part)
                # Minimum rotated rectangle: short side = perpendicular deviation.
                # Long thin boundary-shift slivers (e.g. 2.6 m × 5 km) are correctly
                # measured by their width, not their length.
                mrr = part_utm.minimum_rotated_rectangle
                if mrr.geom_type == "Polygon":
                    c    = list(mrr.exterior.coords)
                    d1   = math.hypot(c[1][0] - c[0][0], c[1][1] - c[0][1])
                    d2   = math.hypot(c[2][0] - c[1][0], c[2][1] - c[1][1])
                    short_side = min(d1, d2)
                    long_side  = max(d1, d2)
                    max_dim_m  = short_side
                    aspect_ratio = (long_side / short_side) if short_side > 0 else 1.0
                else:
                    b = part_utm.bounds
                    max_dim_m = max(b[2] - b[0], b[3] - b[1])
            except Exception:
                pass

            # Skip road-parallel boundary-shift artefacts only when ALL THREE hold:
            #   1. Elongated  : aspect_ratio > 4      (shape is linear)
            #   2. Narrow     : width < 30 m           (thin — road width)
            #   3. Long       : long_side > 150 m      (runs along a road segment)
            #
            # Short narrow features (interior boundary deviations, small protrusions)
            # have long_side < 150 m so they pass through and are kept as real errors.
            #
            # Examples:
            #   5 m × 500 m road sliver  → ratio=100, width=5 m,  length=500 m → skipped ✓
            #   15 m × 200 m road sliver → ratio=13,  width=15 m, length=200 m → skipped ✓
            #   10 m × 60 m  interior error → ratio=6, width=10 m, length=60 m  → kept   ✓
            #   20 m × 80 m  interior error → ratio=4, width=20 m, length=80 m  → kept   ✓
            #   50 m × 300 m real error     → ratio=6, width=50 m               → kept   ✓
            if (aspect_ratio > _GEOM_LINEAR_ASPECT_RATIO
                    and max_dim_m  < _GEOM_ROAD_WIDTH_MAX_M
                    and long_side  > _GEOM_SLIVER_MIN_LENGTH_M):
                continue

            cen = part.centroid
            rec = {_mds_id: fid}
            for col in mds_attr_cols:
                rec[f"mds_{col}"] = mds_attrs.get(col)
            for col in src_attr_cols:
                rec[f"src_{col}"] = src_attrs.get(col)
            rec["_deviation_m"] = round(max_dim_m, 2)
            rec["X"]            = round(cen.x, 9)
            rec["Y"]            = round(cen.y, 9)
            rec["SHAPE_Area"]   = round(part.area, 12)
            rec["SHAPE_Length"] = round(part.length, 12)
            rec["geometry"]     = part

            if max_dim_m < _GEOM_NOISE_FILTER_M:
                lt1_rows.append(rec)
            elif max_dim_m <= eff_threshold:
                lt_buf_rows.append(rec)
            else:
                gt_buf_rows.append(rec)

    # Deduplicate: shared polygon boundaries produce identical sym-diff parts
    # in the comparison of each adjacent feature — keep only the first occurrence.
    lt1_rows    = _dedup_sym_diff_rows(lt1_rows)
    lt_buf_rows = _dedup_sym_diff_rows(lt_buf_rows)
    gt_buf_rows = _dedup_sym_diff_rows(gt_buf_rows)

    # Build output column order: id, mds_*, src_*, metrics, geometry
    empty_cols = ([_mds_id]
                  + [f"mds_{c}" for c in mds_attr_cols]
                  + [f"src_{c}" for c in src_attr_cols]
                  + ["_deviation_m", "X", "Y", "SHAPE_Area", "SHAPE_Length", "geometry"])
    empty_gdf = gpd.GeoDataFrame(columns=empty_cols, geometry="geometry", crs="EPSG:4326")

    def _make(rows):
        if not rows:
            return empty_gdf.copy()
        gdf = gpd.GeoDataFrame(rows, geometry="geometry",
                               crs="EPSG:4326").reset_index(drop=True)
        cols_order = [c for c in empty_cols if c in gdf.columns]
        return gdf[cols_order]

    return _make(lt1_rows), _make(lt_buf_rows), _make(gt_buf_rows)


# ---------------------------------------------------------------------------
# Step 4  Multipart Polygon Analysis
#
# Algorithm:
#   1. For each MDS feature, find the best-matching Source feature by largest
#      spatial overlap (spatial join — no ID required).
#   2. Transfer Source attributes to the MDS feature row.
#   3. Explode both features into individual polygon parts.
#   4. Greedy 1-to-1 matching: each Source part claims the best unmatched MDS
#      part by largest overlap.  Track matched indices on both sides.
#   5. Unmatched MDS parts  → "Extra in MDS"    (MDS was over-updated)
#      Unmatched Source parts → "Missing in MDS" (Source part not updated into MDS)
#
# Example: Source=3 parts, MDS=6 → 3 extra MDS parts reported.
#          Source=3 parts, MDS=2 → 1 Source part missing in MDS reported.
# Feature pairs where both sides have exactly 1 part are skipped.
# ---------------------------------------------------------------------------

def multipart_analysis(
    source_gdf: gpd.GeoDataFrame,
    mds_gdf:    gpd.GeoDataFrame,
) -> Tuple[gpd.GeoDataFrame, dict]:
    source_gdf = _make_valid_gdf(source_gdf)
    mds_gdf    = _make_valid_gdf(mds_gdf)
    utm_epsg = _best_utm(source_gdf)
    src_utm  = source_gdf.to_crs(epsg=utm_epsg).reset_index(drop=True)
    mds_utm  = mds_gdf.to_crs(epsg=utm_epsg).reset_index(drop=True)

    src_mp_count = int((src_utm.geometry.geom_type == "MultiPolygon").sum())
    mds_mp_count = int((mds_utm.geometry.geom_type == "MultiPolygon").sum())

    src_attr_cols = [c for c in src_utm.columns if c != "geometry"]
    mds_attr_cols = [c for c in mds_utm.columns if c != "geometry"]

    src_geoms = src_utm.geometry.values
    src_tree  = STRtree(src_geoms)

    fallout_rows = []

    for mds_idx in range(len(mds_utm)):
        mds_row  = mds_utm.iloc[mds_idx]
        mds_geom = mds_row.geometry

        # Explode MDS feature into individual parts
        mds_parts      = (list(mds_geom.geoms)
                          if mds_geom.geom_type == "MultiPolygon"
                          else [mds_geom])
        mds_part_count = len(mds_parts)

        # ── Spatial join: find best-matching Source feature by overlap ────────
        cand_idxs    = src_tree.query(mds_geom)
        best_src_idx = None
        best_overlap = 0.0
        for ci in cand_idxs:
            sg = src_geoms[ci]
            if not mds_geom.intersects(sg):
                continue
            try:
                ov = mds_geom.intersection(sg).area
            except Exception:
                ov = 0.0
            if ov > best_overlap:
                best_overlap = ov
                best_src_idx = ci

        # ── No Source match → all MDS parts are fallout (skip if single) ─────
        if best_src_idx is None:
            if mds_part_count == 1:
                continue   # single unmatched feature — not a multipart issue
            for part in mds_parts:
                cen = part.centroid
                rec = {f"mds_{c}": mds_row[c] for c in mds_attr_cols}
                rec.update({f"src_{c}": None for c in src_attr_cols})
                rec["_src_part_count"] = 0
                rec["_mds_part_count"] = mds_part_count
                rec["_reason"]         = "No matching Source feature found"
                rec["X"]               = round(cen.x, 9)
                rec["Y"]               = round(cen.y, 9)
                rec["SHAPE_Area"]      = round(part.area, 2)
                rec["SHAPE_Length"]    = round(part.length, 2)
                rec["geometry"]        = part
                fallout_rows.append(rec)
            continue

        src_row   = src_utm.iloc[best_src_idx]
        src_geom  = src_row.geometry
        src_parts = (list(src_geom.geoms)
                     if src_geom.geom_type == "MultiPolygon"
                     else [src_geom])
        src_part_count = len(src_parts)

        # Skip single↔single pairs — no multipart concern
        if src_part_count == 1 and mds_part_count == 1:
            continue

        # ── Greedy 1-to-1 matching: Source parts claim MDS parts ─────────────
        mds_parts_tree  = STRtree(mds_parts)
        matched_mds_idx = set()
        matched_src_idx = set()

        for si, sp in enumerate(src_parts):
            cands   = mds_parts_tree.query(sp)
            best_mi = None
            best_ov = 0.0
            for mi in cands:
                if mi in matched_mds_idx:
                    continue
                mp = mds_parts[mi]
                if not sp.intersects(mp):
                    continue
                try:
                    ov = sp.intersection(mp).area
                except Exception:
                    ov = 0.0
                if ov > best_ov:
                    best_ov = ov
                    best_mi = mi
            if best_mi is not None:
                matched_mds_idx.add(best_mi)
                matched_src_idx.add(si)

        # ── Extra MDS parts (not in Source) → fallout ────────────────────────
        for mi, mp in enumerate(mds_parts):
            if mi in matched_mds_idx:
                continue
            cen = mp.centroid
            rec = {f"mds_{c}": mds_row[c] for c in mds_attr_cols}
            rec.update({f"src_{c}": src_row[c] for c in src_attr_cols})
            rec["_src_part_count"] = src_part_count
            rec["_mds_part_count"] = mds_part_count
            rec["_issue_type"]     = "Extra in MDS"
            rec["_reason"]         = (
                f"MDS part has no matching Source part — "
                f"Source has {src_part_count} part(s), MDS has {mds_part_count} part(s)"
            )
            rec["X"]            = round(cen.x, 9)
            rec["Y"]            = round(cen.y, 9)
            rec["SHAPE_Area"]   = round(mp.area, 2)
            rec["SHAPE_Length"] = round(mp.length, 2)
            rec["geometry"]     = mp
            fallout_rows.append(rec)

        # ── Source parts missing in MDS → fallout ────────────────────────────
        for si, sp in enumerate(src_parts):
            if si in matched_src_idx:
                continue
            cen = sp.centroid
            rec = {f"mds_{c}": mds_row[c] for c in mds_attr_cols}
            rec.update({f"src_{c}": src_row[c] for c in src_attr_cols})
            rec["_src_part_count"] = src_part_count
            rec["_mds_part_count"] = mds_part_count
            rec["_issue_type"]     = "Missing in MDS"
            rec["_reason"]         = (
                f"Source part not found in MDS — "
                f"Source has {src_part_count} part(s), MDS has {mds_part_count} part(s)"
            )
            rec["X"]            = round(cen.x, 9)
            rec["Y"]            = round(cen.y, 9)
            rec["SHAPE_Area"]   = round(sp.area, 2)
            rec["SHAPE_Length"] = round(sp.length, 2)
            rec["geometry"]     = sp
            fallout_rows.append(rec)

    if fallout_rows:
        fallout_gdf = (
            gpd.GeoDataFrame(fallout_rows, geometry="geometry", crs=mds_utm.crs)
            .to_crs(epsg=4326)
            .reset_index(drop=True)
        )
        fallout_gdf = _reorder_cols(fallout_gdf)
    else:
        fallout_gdf = mds_gdf.iloc[0:0].copy()

    extra_count   = sum(1 for r in fallout_rows if r.get("_issue_type") == "Extra in MDS")
    missing_count = sum(1 for r in fallout_rows if r.get("_issue_type") == "Missing in MDS")

    return fallout_gdf, {
        "src_multipart_count":         src_mp_count,
        "mds_multipart_count":         mds_mp_count,
        "discrepancy_count":           len(fallout_rows),
        "multipart_extra_count":       extra_count,
        "multipart_missing_count":     missing_count,
    }


# ---------------------------------------------------------------------------
# Step 5  Polygon Quality Check  (symmetric difference, 5 m hardcoded filter)
#
# Logic:
#   1. Match each MDS polygon to its Source polygon by shared ID (uuid).
#   2. Compute symmetric_difference(source_geom, mds_geom).
#   3. Decompose the sym-diff into individual parts.
#   4. Project each part to UTM; measure bounding-box max dimension.
#      Parts ≤ 5 m  → within tolerance, skip (matched OK).
#      Parts  > 5 m → error, output one row.
#   5. Each output row carries both mds_* and src_* attributes.
#
# The 5 m threshold is intentionally separate from the Step 3 buffer distance
# and is hardcoded (_POLY_QC_DEVIATION_M).
#
# Output: POLYGON layer "Polygon_Quality_Check"
# ---------------------------------------------------------------------------

_POLY_QC_DEVIATION_M = 5.0   # hardcoded — do not expose in UI

def _find_id_col(mds_gdf: gpd.GeoDataFrame,
                 src_gdf: Optional[gpd.GeoDataFrame] = None,
                 candidates=("uuid", "feat_id", "UUID", "FEAT_ID", "CP5", "cp5")) -> Optional[str]:
    """
    Find a shared ID column between MDS and Source (same column name in both).
    Priority: candidates list → any column with 'uuid' in name → any common column.
    If src_gdf is None, only MDS columns are searched.
    """
    src_cols = set(src_gdf.columns) if src_gdf is not None else set(mds_gdf.columns)

    # 1. Exact candidate match in both
    for c in candidates:
        if c in mds_gdf.columns and c in src_cols:
            return c

    # 2. Any column whose name contains 'uuid', present in both
    for c in mds_gdf.columns:
        if "uuid" in c.lower() and c in src_cols:
            return c

    # 3. Any non-geometry column common to both datasets
    if src_gdf is not None:
        mds_non_geom = {c for c in mds_gdf.columns if c != "geometry"}
        src_non_geom = {c for c in src_gdf.columns  if c != "geometry"}
        common = mds_non_geom & src_non_geom
        if common:
            return sorted(common)[0]   # deterministic choice

    return None


def _find_id_col_pair(
    src_gdf:      gpd.GeoDataFrame,
    mds_gdf:      gpd.GeoDataFrame,
    src_id_col:   Optional[str] = None,
    mds_id_col:   Optional[str] = None,
    candidates=("uuid", "feat_id", "UUID", "FEAT_ID", "CP5", "cp5"),
    value_overlap_threshold: float = 0.5,
) -> Tuple[Optional[str], Optional[str]]:
    """Return (src_id_col, mds_id_col) — the columns to use for matching.

    Resolution order:
    1. Explicit overrides (src_id_col / mds_id_col) — used when caller specifies them.
    2. Same-name match via _find_id_col (existing logic).
    3. Value-overlap matching: find a Source column and MDS column whose values
       overlap significantly (>= value_overlap_threshold), even if names differ.
       This handles cases like SourcePost (Source) ↔ CP5 (MDS).
    Returns (None, None) if no match is found.
    """
    # 1. Explicit overrides
    if src_id_col and mds_id_col:
        if src_id_col in src_gdf.columns and mds_id_col in mds_gdf.columns:
            return src_id_col, mds_id_col

    # 2. Same-name match (existing behaviour)
    same = _find_id_col(mds_gdf, src_gdf=src_gdf, candidates=candidates)
    if same is not None:
        return same, same

    # 3. Value-overlap matching (handles different column names for the same concept)
    # Only consider columns with enough distinct values to be meaningful IDs.
    # Skip known auto-increment / row-index names to avoid false positives.
    _MIN_DISTINCT = 5
    _SKIP_COLS = {"fid", "fid_1", "objectid", "object_id", "oid", "gid",
                  "id", "row_id", "rowid", "index"}
    mds_cols = [c for c in mds_gdf.columns
                if c != "geometry" and c.lower() not in _SKIP_COLS]
    src_cols  = [c for c in src_gdf.columns
                 if c != "geometry" and c.lower() not in _SKIP_COLS]

    best_overlap = 0.0
    best_pair    = (None, None)

    for mc in mds_cols:
        mds_vals = set(mds_gdf[mc].dropna().astype(str))
        if len(mds_vals) < _MIN_DISTINCT:
            continue
        for sc in src_cols:
            src_vals = set(src_gdf[sc].dropna().astype(str))
            if len(src_vals) < _MIN_DISTINCT:
                continue
            overlap_n = len(mds_vals & src_vals)
            if overlap_n == 0:
                continue
            ratio = overlap_n / max(len(mds_vals), len(src_vals))
            if ratio > best_overlap:
                best_overlap = ratio
                best_pair    = (sc, mc)

    if best_overlap >= value_overlap_threshold:
        import warnings
        warnings.warn(
            f"Step 3 ID column auto-matched by value overlap "
            f"({best_overlap:.0%}): Source.{best_pair[0]} ↔ MDS.{best_pair[1]}"
        )
        return best_pair

    return None, None


def polygon_quality_check(
    source_gdf:    gpd.GeoDataFrame,
    mds_gdf:       gpd.GeoDataFrame,
    id_col:        Optional[str] = None,
    progress_callback=None,
) -> gpd.GeoDataFrame:
    """
    For every MDS polygon matched to a Source polygon (by uuid):
      1. Compute symmetric_difference(source, mds).
      2. Decompose into individual parts.
      3. Project each part to UTM; measure bounding-box max dimension.
         Part ≤ _POLY_QC_DEVIATION_M (5 m) → within tolerance, skip.
         Part  > 5 m → error, output one row per failing part.
      4. Each row carries both mds_* and src_* attributes.
    """
    from pyproj import Transformer
    from shapely.ops import transform as shp_transform

    src = source_gdf.to_crs(epsg=4326).copy()
    mds = mds_gdf.to_crs(epsg=4326).copy()

    # ── Detect shared ID column ───────────────────────────────────────────────
    if id_col is None:
        id_col = _find_id_col(mds, src_gdf=src)

    if id_col is None or id_col not in src.columns or id_col not in mds.columns:
        # No common ID column — cannot do UUID-based polygon QC.
        # Return an empty result rather than crashing the whole run.
        import warnings
        warnings.warn(
            f"Polygon Quality Check skipped: no common ID column between "
            f"Source {list(src.columns)} and MDS {list(mds.columns)}."
        )
        return gpd.GeoDataFrame(columns=["geometry"], geometry="geometry",
                                crs="EPSG:4326")

    if progress_callback:
        progress_callback("Polygon Quality: indexing matched pairs…")

    src_ids    = set(src[id_col].dropna().astype(str))
    mds_ids    = set(mds[id_col].dropna().astype(str))
    common_ids = src_ids & mds_ids

    src_idx = (src[src[id_col].astype(str).isin(common_ids)].set_index(id_col))
    mds_idx = (mds[mds[id_col].astype(str).isin(common_ids)].set_index(id_col))

    mds_attr_cols = [c for c in mds.columns if c not in ("geometry", id_col)]
    src_attr_cols = [c for c in src.columns if c not in ("geometry", id_col)]

    if progress_callback:
        progress_callback("Polygon Quality: computing symmetric differences…")

    # Cache UTM Transformer objects by EPSG zone (avoid per-feature creation)
    _t_cache: dict = {}

    def _get_t(epsg: int):
        if epsg not in _t_cache:
            _t_cache[epsg] = Transformer.from_crs(
                "EPSG:4326", f"EPSG:{epsg}", always_xy=True)
        return _t_cache[epsg]

    # Deduplicate index (ensures .loc returns Series, not DataFrame slice)
    src_idx = src_idx[~src_idx.index.duplicated(keep="first")]
    mds_idx = mds_idx[~mds_idx.index.duplicated(keep="first")]

    # Pre-build geometry and attribute dicts for O(1) lookup per feature ID.
    # DataFrame .loc[] on string index is O(log N); dict lookup is O(1).
    src_geom_map  = dict(zip(src_idx.index.astype(str), src_idx["geometry"]))
    mds_geom_map  = dict(zip(mds_idx.index.astype(str), mds_idx["geometry"]))
    src_attrs_map = src_idx[src_attr_cols].to_dict(orient="index")
    mds_attrs_map = mds_idx[mds_attr_cols].to_dict(orient="index")

    diff_rows = []
    n = len(common_ids)

    for i, fid in enumerate(common_ids):
        if progress_callback and i % max(1, n // 20) == 0:
            progress_callback("Polygon Quality: comparing geometries…")

        src_geom = src_geom_map.get(fid)
        mds_geom = mds_geom_map.get(fid)

        if src_geom is None or mds_geom is None:
            continue
        if src_geom.equals(mds_geom):
            continue   # identical — no issue

        try:
            if not src_geom.is_valid:
                src_geom = src_geom.buffer(0)
            if not mds_geom.is_valid:
                mds_geom = mds_geom.buffer(0)
            sym_diff = src_geom.symmetric_difference(mds_geom)
        except Exception:
            try:
                from shapely.validation import make_valid
                src_geom = make_valid(src_geom)
                mds_geom = make_valid(mds_geom)
                sym_diff = src_geom.symmetric_difference(mds_geom)
            except Exception:
                continue
        if sym_diff is None or sym_diff.is_empty:
            continue

        # ── UTM transformer for this pair ─────────────────────────────────────
        utm_epsg = _geom_utm_epsg(src_geom)
        try:
            project = _get_t(utm_epsg).transform   # cached by zone
        except Exception:
            project = None   # fallback: skip the distance filter

        # ── Decompose sym-diff into individual parts ──────────────────────────
        if sym_diff.geom_type == "Polygon":
            parts = [sym_diff]
        elif sym_diff.geom_type in ("MultiPolygon", "GeometryCollection"):
            parts = [g for g in sym_diff.geoms
                     if g.geom_type == "Polygon" and not g.is_empty]
        else:
            parts = []

        mds_row = mds_attrs_map.get(fid, {})
        src_row = src_attrs_map.get(fid, {})

        for part in parts:
            # ── 5 m deviation filter (hardcoded) ─────────────────────────────
            if project is not None:
                try:
                    part_utm = shp_transform(project, part)
                    b = part_utm.bounds                    # (minx, miny, maxx, maxy)
                    max_dim = max(b[2] - b[0], b[3] - b[1])
                    if max_dim <= _POLY_QC_DEVIATION_M:
                        continue   # within 5 m tolerance → not an error
                except Exception:
                    pass           # projection failed — include the part anyway

            # ── Build output row ──────────────────────────────────────────────
            cen = part.centroid
            rec = {id_col: fid}

            for c in mds_attr_cols:
                rec[f"mds_{c}"] = mds_row.get(c)
            for c in src_attr_cols:
                rec[f"src_{c}"] = src_row.get(c)

            rec["mds_area"]     = round(mds_geom.area,  12)
            rec["src_area"]     = round(src_geom.area,  12)
            rec["area_diff"]    = round(abs(mds_geom.area - src_geom.area), 12)
            rec["symdiff_area"] = round(sym_diff.area,  12)
            rec["X"]            = round(cen.x,  9)
            rec["Y"]            = round(cen.y,  9)
            rec["SHAPE_Area"]   = round(part.area,   12)
            rec["SHAPE_Length"] = round(part.length, 12)
            rec["geometry"]     = part
            diff_rows.append(rec)

    # ── Build output GDF ──────────────────────────────────────────────────────
    empty_mds  = [f"mds_{c}" for c in mds_attr_cols]
    empty_src  = [f"src_{c}" for c in src_attr_cols]
    empty_cols = ([id_col] + empty_mds + empty_src +
                  ["mds_area", "src_area", "area_diff", "symdiff_area",
                   "X", "Y", "SHAPE_Area", "SHAPE_Length", "geometry"])

    if not diff_rows:
        return gpd.GeoDataFrame(columns=empty_cols, geometry="geometry",
                                crs="EPSG:4326")

    out = gpd.GeoDataFrame(diff_rows, geometry="geometry",
                           crs="EPSG:4326").reset_index(drop=True)

    lead = [id_col]
    for base in ("uuid", "CP5", "ddctType"):
        mds_col = f"mds_{base}"
        if mds_col in out.columns and mds_col not in lead:
            lead.append(mds_col)
    rest = [c for c in out.columns if c not in lead + ["geometry"]]
    return out[lead + rest + ["geometry"]]


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_qc(
    source_gdb:       str,
    mds_gdb:          str,
    source_layer:     Optional[str] = None,
    mds_layer:        Optional[str] = None,
    spike_angle:      float = 15.0,
    buffer_dist_m:    float = 200.0,
    run_spike:        bool = True,
    run_sliver:       bool = True,
    run_buffer:       bool = True,
    run_multipart:    bool = True,
    run_poly_quality: bool = True,
    progress_callback=None,
    source_id_col:    Optional[str] = None,
    mds_id_col:       Optional[str] = None,
) -> QCResult:
    # If nothing selected, run everything
    if not any([run_spike, run_sliver, run_buffer, run_multipart, run_poly_quality]):
        run_spike = run_sliver = run_buffer = run_multipart = run_poly_quality = True

    timestamp   = datetime.datetime.now().strftime("%Y%m%d")
    active      = sum([run_spike, run_sliver, run_buffer, run_multipart, run_poly_quality])
    total_steps = 2 + active   # 2 reads + active checks

    step = [0]
    def _cb(msg):
        if progress_callback:
            progress_callback(step[0], total_steps, msg)
        step[0] += 1

    _cb("Reading Source dataset...")
    source_gdf = read_layer(source_gdb, source_layer)

    _cb("Reading MDS dataset...")
    mds_gdf = read_layer(mds_gdb, mds_layer)

    _empty_point   = gpd.GeoDataFrame(columns=["geometry"], geometry="geometry", crs="EPSG:4326")
    _empty_polygon = gpd.GeoDataFrame(columns=["geometry"], geometry="geometry", crs="EPSG:4326")
    _empty_line    = gpd.GeoDataFrame(columns=["geometry"], geometry="geometry", crs="EPSG:4326")

    # Step 1 — Spike Detection
    if run_spike:
        _cb("Step 1 - Spike Detection (angle ≤ {} deg)...".format(spike_angle))
        spike_gdf = detect_spikes(mds_gdf, spike_angle=spike_angle)
    else:
        spike_gdf = _empty_point.copy()

    # Step 2 — Sliver Detection
    if run_sliver:
        _cb("Step 2 - Sliver Detection (area rounds to 0)...")
        sliver_gdf = detect_slivers(mds_gdf)
    else:
        sliver_gdf = _empty_polygon.copy()

    # Step 3 — Geometry Failure Check (sym-diff, configurable threshold)
    eff_t = _effective_threshold(buffer_dist_m)
    # Resolve ID columns now (before Step 3) so the result goes into summary
    _resolved_src_id, _resolved_mds_id = _find_id_col_pair(
        source_gdf.to_crs(epsg=4326),
        mds_gdf.to_crs(epsg=4326),
        src_id_col=source_id_col,
        mds_id_col=mds_id_col,
    )
    if run_buffer:
        _cb(f"Step 3 - Geometry Failure Check (buffer {buffer_dist_m} m, threshold {eff_t} m)...")
        geom_lt1_gdf, geom_lt_buf_gdf, geom_gt_buf_gdf = buffer_comparison(
            source_gdf, mds_gdf, buffer_dist_m=buffer_dist_m,
            src_id_col=_resolved_src_id, mds_id_col=_resolved_mds_id)
    else:
        geom_lt1_gdf = geom_lt_buf_gdf = geom_gt_buf_gdf = _empty_polygon.copy()

    # Step 4 — Multipart Analysis
    if run_multipart:
        _cb("Step 4 - Multipart Polygon Analysis...")
        multipart_gdf, mp_stats = multipart_analysis(source_gdf, mds_gdf)
    else:
        multipart_gdf = _empty_polygon.copy()
        mp_stats = {"src_multipart_count": 0, "mds_multipart_count": 0, "discrepancy_count": 0}

    # Step 5 — Polygon Quality Check (sym-diff, 5 m hardcoded filter)
    if run_poly_quality:
        _cb("Step 5 - Polygon Quality Check (sym-diff, 5 m tolerance)...")
        poly_quality_gdf = polygon_quality_check(source_gdf, mds_gdf)
    else:
        poly_quality_gdf = _empty_polygon.copy()

    summary = {
        "source_feature_count":        len(source_gdf),
        "mds_feature_count":           len(mds_gdf),
        "spike_count":                 len(spike_gdf),
        "sliver_count":                len(sliver_gdf),
        "geom_lt1_count":              len(geom_lt1_gdf),
        "geom_lt_buf_count":           len(geom_lt_buf_gdf),
        "geom_gt_buf_count":           len(geom_gt_buf_gdf),
        "buffer_dist_m":               buffer_dist_m,
        "effective_threshold_m":       eff_t,
        "src_multipart_count":         mp_stats["src_multipart_count"],
        "mds_multipart_count":         mp_stats["mds_multipart_count"],
        "multipart_discrepancy_count": mp_stats["discrepancy_count"],
        "multipart_extra_count":       mp_stats.get("multipart_extra_count", 0),
        "multipart_missing_count":     mp_stats.get("multipart_missing_count", 0),
        "poly_quality_count":          len(poly_quality_gdf),
        "spike_angle_threshold":       spike_angle,
        "checks_run": {
            "spike":        run_spike,
            "sliver":       run_sliver,
            "buffer":       run_buffer,
            "multipart":    run_multipart,
            "poly_quality": run_poly_quality,
        },
        "step3_src_id_col": _resolved_src_id,
        "step3_mds_id_col": _resolved_mds_id,
        "timestamp": timestamp,
    }

    return QCResult(
        spike            = spike_gdf,
        sliver           = sliver_gdf,
        geom_lt1         = geom_lt1_gdf,
        geom_lt_buf      = geom_lt_buf_gdf,
        geom_gt_buf      = geom_gt_buf_gdf,
        multipart        = multipart_gdf,
        poly_quality     = poly_quality_gdf,
        summary          = summary,
        timestamp        = timestamp,
    )


# ---------------------------------------------------------------------------
# Multi-layer result merging
# ---------------------------------------------------------------------------

def merge_qc_results(results: List["QCResult"]) -> "QCResult":
    """Merge a list of per-layer QCResult objects into one combined result.

    All GeoDataFrames are concatenated (reset index).  Summary counts are
    summed.  The effective_threshold_m and other non-count settings are taken
    from the first result (they should be identical across all layers).
    """
    if not results:
        raise ValueError("merge_qc_results: empty results list")
    if len(results) == 1:
        return results[0]

    def _concat(gdfs: list) -> gpd.GeoDataFrame:
        non_empty = [g for g in gdfs if g is not None and len(g) > 0]
        if not non_empty:
            return gpd.GeoDataFrame()
        import pandas as pd
        return gpd.GeoDataFrame(
            pd.concat(non_empty, ignore_index=True),
            crs=non_empty[0].crs,
        )

    # Merge summary: sum all *_count keys; keep scalars from first result.
    merged_summary = dict(results[0].summary)
    count_keys = [k for k in merged_summary if k.endswith("_count")]
    for k in count_keys:
        merged_summary[k] = sum(r.summary.get(k, 0) for r in results)
    # discrepancy_count is derived — recalculate
    merged_summary["multipart_discrepancy_count"] = (
        merged_summary.get("multipart_extra_count", 0)
        + merged_summary.get("multipart_missing_count", 0)
    )
    # Use timestamp from the last (most recent) run
    merged_summary["timestamp"] = results[-1].timestamp

    return QCResult(
        spike=_concat([r.spike        for r in results]),
        sliver=_concat([r.sliver      for r in results]),
        geom_lt1=_concat([r.geom_lt1  for r in results]),
        geom_lt_buf=_concat([r.geom_lt_buf for r in results]),
        geom_gt_buf=_concat([r.geom_gt_buf for r in results]),
        multipart=_concat([r.multipart for r in results]),
        poly_quality=_concat([r.poly_quality for r in results]),
        summary=merged_summary,
        timestamp=results[-1].timestamp,
    )


# ---------------------------------------------------------------------------
# Output writers
# ---------------------------------------------------------------------------

def _get_layers(result: "QCResult"):
    """Return (layer_name, getter, check_key) tuples; Step 3 names are dynamic.

    check_key matches the keys in result.summary["checks_run"].
    Callers should skip any entry whose check_key maps to False in checks_run.
    """
    t = result.summary.get("effective_threshold_m", 200.0)
    lt_name, gt_name = _step3_layer_names(t)
    return [
        ("Spike_Output",                  lambda r: r.spike,        "spike"),
        ("Sliver_Suspicious_Area_Output", lambda r: r.sliver,       "sliver"),
        ("Geometry_Failed_Less_Than_1m",  lambda r: r.geom_lt1,     "buffer"),
        (lt_name,                         lambda r: r.geom_lt_buf,  "buffer"),
        (gt_name,                         lambda r: r.geom_gt_buf,  "buffer"),
        ("Multipart_Validation",          lambda r: r.multipart,    "multipart"),
        ("Polygon_Quality_Check",         lambda r: r.poly_quality, "poly_quality"),
    ]


def _check_enabled(result: "QCResult", check_key: str) -> bool:
    """Return True if the given check was enabled for this run (default True)."""
    return result.summary.get("checks_run", {}).get(check_key, True)


def _write_layers(result: QCResult, output_path: str, driver: str) -> str:
    import shutil
    # Remove any existing output so we start clean
    if driver == "OpenFileGDB" and os.path.exists(output_path):
        shutil.rmtree(output_path)
    elif os.path.exists(output_path):
        os.remove(output_path)

    written = 0
    for layer_name, getter, check_key in _get_layers(result):
        if not _check_enabled(result, check_key):
            continue          # omit layers whose check was disabled
        gdf = getter(result)
        if gdf is None or len(gdf) == 0:
            gdf = gpd.GeoDataFrame({"geometry": []}, crs="EPSG:4326")
        # GPKG: first layer creates the file (mode='w'), rest append (mode='a')
        # OpenFileGDB: mode param is ignored by fiona — file was already removed above
        mode = "a" if (driver == "GPKG" and written > 0) else "w"
        gdf.to_file(output_path, layer=layer_name, driver=driver, mode=mode)
        written += 1

    return output_path


def write_gdb(result: QCResult, output_path: str) -> str:
    import fiona
    writable = fiona.supported_drivers.get("OpenFileGDB", "") in ("rw", "w")
    if not writable:
        raise RuntimeError(
            "OpenFileGDB write driver not available. "
            "Run from ArcGIS Pro Python, or GDAL >= 3.6.")
    return _write_layers(result, output_path, "OpenFileGDB")


def write_gpkg(result: QCResult, output_path: str) -> str:
    return _write_layers(result, output_path, "GPKG")


def write_shp(result: QCResult, output_dir: str) -> str:
    """Write each QC layer as a separate Shapefile inside output_dir."""
    os.makedirs(output_dir, exist_ok=True)
    for layer_name, getter, check_key in _get_layers(result):
        if not _check_enabled(result, check_key):
            continue          # omit layers whose check was disabled
        gdf = getter(result)
        out_path = os.path.join(output_dir, layer_name + ".shp")
        if gdf is not None and len(gdf) > 0:
            # Shapefile column names limited to 10 chars
            rename_map = {}
            seen = {"geometry"}
            for col in gdf.columns:
                if col == "geometry":
                    continue
                new_col = col[:10]
                base, suffix = new_col, 0
                while new_col in seen:
                    suffix += 1
                    new_col = (base[:9] + str(suffix))[:10]
                seen.add(new_col)
                if new_col != col:
                    rename_map[col] = new_col
            out_gdf = gdf.rename(columns=rename_map) if rename_map else gdf
            out_gdf.to_file(out_path, driver="ESRI Shapefile")
        else:
            gpd.GeoDataFrame({"geometry": []}, crs="EPSG:4326").to_file(
                out_path, driver="ESRI Shapefile")
    return output_dir


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

_EXCEL_MAX_ROWS = 1_048_575   # Excel hard limit (1,048,576 rows incl. header)


def to_excel_bytes(result: QCResult) -> bytes:
    """
    Serialise QCResult to an Excel workbook (.xlsx) in memory.

    Large layers (> _EXCEL_MAX_ROWS) are automatically split into
    SheetName_Pt1 / SheetName_Pt2 / ... so no data is lost.
    Excel sheet names are capped at 31 characters.
    """
    import io
    from openpyxl.styles import Font, PatternFill, Alignment

    def _prep(gdf):
        return pd.DataFrame(gdf.drop(columns=["geometry"], errors="ignore"))

    def _style(ws):
        hfill = PatternFill("solid", fgColor="1A56DB")
        hfont = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill      = hfill
            cell.font      = hfont
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    def _write(writer, base_name, df):
        """Write df, splitting into _Pt1/_Pt2/... sheets if rows exceed limit."""
        if len(df) <= _EXCEL_MAX_ROWS:
            sname = base_name[:31]
            df.to_excel(writer, sheet_name=sname, index=False)
            _style(writer.sheets[sname])
        else:
            n_parts = math.ceil(len(df) / _EXCEL_MAX_ROWS)
            for p in range(n_parts):
                chunk = df.iloc[p * _EXCEL_MAX_ROWS : (p + 1) * _EXCEL_MAX_ROWS].copy()
                # Build sheet name <= 31 chars: trim base to leave room for suffix
                suffix = f"_Pt{p + 1}"
                sname  = (base_name[: 31 - len(suffix)] + suffix)
                chunk.to_excel(writer, sheet_name=sname, index=False)
                _style(writer.sheets[sname])

    s = result.summary
    summary_df = pd.DataFrame([
        ("Source Feature Count",          s["source_feature_count"],          "-"),
        ("MDS Feature Count",             s["mds_feature_count"],             "-"),
        ("Spike Vertices (Spike_Output)",
         s["spike_count"],
         "MDS angle <= {} deg".format(s["spike_angle_threshold"])),
        ("Sliver Polygons (Sliver_Suspicious_Area_Output)",
         s["sliver_count"],
         "Area rounds to 0 (6 d.p.)"),
        ("Geometry Failed < 1 m (Geometry_Failed_Less_Than_1m)",
         s["geom_lt1_count"],
         "Sym-diff parts < 1 m -- coordinate rounding / positional noise"),
        (
            "Geometry Failed 1-{t} m (Geometry_Failed_Less_Than_{t}m)".format(
                t=int(s["effective_threshold_m"])
            ),
            s["geom_lt_buf_count"],
            "Sym-diff parts: deviation 1-{t} m (minor)".format(
                t=int(s["effective_threshold_m"])
            ),
        ),
        (
            "Geometry Failed >{t} m (Geometry_Failed_Greater_Than_{t}m)".format(
                t=int(s["effective_threshold_m"])
            ),
            s["geom_gt_buf_count"],
            "Sym-diff parts: deviation > {t} m (major)".format(
                t=int(s["effective_threshold_m"])
            ),
        ),
        ("Source Multipart Features",   s["src_multipart_count"],           "-"),
        ("MDS Multipart Features",      s["mds_multipart_count"],           "-"),
        ("Multipart -- Extra in MDS",   s.get("multipart_extra_count", 0),
         "MDS parts with no matching Source part (over-updated)"),
        ("Multipart -- Missing in MDS", s.get("multipart_missing_count", 0),
         "Source parts not found in MDS (update missing)"),
        ("Polygon Quality Check",       s["poly_quality_count"],
         "Sym-diff parts (by UUID) with bounding-box deviation > 5 m"),
        ("Step 3 ID columns used",      "",
         "{} <-> {}".format(
             s.get("step3_src_id_col") or "(not matched)",
             s.get("step3_mds_id_col") or "(not matched)",
         )),
    ], columns=["Check", "Count", "Notes"])

    # Append a note for any checks that were disabled
    checks_run = s.get("checks_run", {})
    _label_map = {
        "spike":        "Spike Detection",
        "sliver":       "Sliver Detection",
        "buffer":       "Failed Geometry",
        "multipart":    "Multipart Check",
        "poly_quality": "Polygon Quality Check",
    }
    skipped = [_label_map[k] for k, v in checks_run.items() if not v]
    if skipped:
        skip_row = pd.DataFrame(
            [("Checks skipped (not run)", ", ".join(skipped), "Output layers omitted")],
            columns=["Check", "Count", "Notes"],
        )
        summary_df = pd.concat([summary_df, skip_row], ignore_index=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _write(writer, "Summary", summary_df)
        for layer_name, getter, check_key in _get_layers(result):
            if not _check_enabled(result, check_key):
                continue
            gdf = getter(result)
            if gdf is None or len(gdf) == 0:
                continue
            _write(writer, layer_name, _prep(gdf))
    return buf.getvalue()
