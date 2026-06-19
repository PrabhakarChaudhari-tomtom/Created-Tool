# =============================================================================
# Copyright (c) 2025 ADP Team, TomTom. All rights reserved.
# This file is proprietary and confidential. Developed exclusively by the
# ADP Team at TomTom. Unauthorised copying, distribution or modification
# of this file, via any medium, is strictly prohibited.
# Contact: prabhakar.chaudhari@tomtom.com | sachin.shete@tomtom.com
# =============================================================================
"""
AA1 Name Checker -- Core Analysis Logic

Process (mirrors FME Workspace Runner + AA1_Name_Analysis_V1):
  1. Query information_schema.schemata on BOTH MNR servers to discover all
     schemas that start with the OLD version prefix.
       Server 001: EUR + CIS schemas
       Server 002: Americas, APAC, Africa, ME schemas
  2. For each OLD schema, derive the matching NEW schema (same suffix).
     Check BOTH servers for the new schema.
  3. Query mnr_admin_area (feat_type=1112, artificial=false) -- name + official_code.
  4. Join on feat_id, compare name_old vs name_new AND OldOfficialCode vs NewOfficialCode.
  5. Route to three output DataFrames:
       matched          -> All_Matched_AA1   (name AND official_code both match)
       name_error       -> Name_Error        (name OR official_code differs)
       missing_feature  -> Missing_feature   (feat_id absent from new schema)
"""

from __future__ import annotations

import datetime
from dataclasses import dataclass, field
from typing import Optional, List, Tuple, Dict

import pandas as pd
import psycopg2


# -- Output column order (matches FME Excel output) ---------------------------
OUTPUT_COLS = [
    "feat_id",
    "artificial",
    "feat_type",
    "name_old",
    "standard_lang_new",
    "country_code",
    "name_new",
    "OldOfficialCode",
    "NewOfficialCode",
    "_timestamp",
]


@dataclass
class DBConfig:
    host: str
    port: int
    dbname: str
    user: str
    password: str


@dataclass
class AnalysisResult:
    matched: pd.DataFrame
    name_error: pd.DataFrame
    missing_feature: pd.DataFrame
    timestamp: str
    schemas_processed: List[Tuple[str, str]] = field(default_factory=list)
    schemas_new_missing: List[str] = field(default_factory=list)


def _connect(cfg: DBConfig):
    return psycopg2.connect(
        host=cfg.host,
        port=cfg.port,
        dbname=cfg.dbname,
        user=cfg.user,
        password=cfg.password,
        connect_timeout=15,
    )


def _get_all_schemas(cfg: DBConfig) -> set:
    """Return set of all schema names on one server."""
    conn = _connect(cfg)
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT schema_name FROM information_schema.schemata ORDER BY schema_name"
        )
        return {row[0] for row in cur.fetchall()}
    finally:
        conn.close()


def discover_schema_pairs_multi(
    configs: List[DBConfig],
    prefix_old: str,
    prefix_new: str,
    country_filter: Optional[str] = None,
) -> Tuple[
    List[Tuple[str, DBConfig, str, DBConfig]],
    List[Tuple[str, DBConfig]],
]:
    """
    Query information_schema.schemata on ALL servers and build schema pairs.
    Returns:
      pairs       -- (old_schema, old_cfg, new_schema, new_cfg)
      missing_new -- (old_schema, old_cfg)  where new schema does not exist
    """
    # Schemas that the FME workspace explicitly excludes (Tester_5 filter).
    # These are duplicate schemas that exist alongside the primary country schemas
    # and would cause double-counting if processed.  Suffixes are case-insensitive.
    _EXCLUDED_SUFFIXES = {"sea_ind_ind", "sea_chn_chn", "mea_isr_isr"}

    schema_to_cfg: Dict[str, DBConfig] = {}
    for cfg in configs:
        try:
            schemas = _get_all_schemas(cfg)
            for s in schemas:
                if s not in schema_to_cfg:
                    schema_to_cfg[s] = cfg
        except Exception:
            pass

    all_schema_names_lower = {s.lower(): s for s in schema_to_cfg}
    prefix_old_lower = prefix_old.lower()
    cf = country_filter.lower() if country_filter else None

    pairs: List[Tuple[str, DBConfig, str, DBConfig]] = []
    missing_new: List[Tuple[str, DBConfig]] = []

    for schema in sorted(schema_to_cfg.keys()):
        schema_lower = schema.lower()
        if not schema_lower.startswith(prefix_old_lower):
            continue

        suffix = schema[len(prefix_old):]

        # Skip schemas that FME excludes to prevent duplicate country data
        suffix_tail = suffix.lower().lstrip("_")
        if any(suffix_tail.endswith(excl) for excl in _EXCLUDED_SUFFIXES):
            continue

        if cf and cf not in suffix.lower():
            continue

        old_cfg = schema_to_cfg[schema]
        new_schema = prefix_new + suffix

        if new_schema in schema_to_cfg:
            pairs.append((schema, old_cfg, new_schema, schema_to_cfg[new_schema]))
        elif new_schema.lower() in all_schema_names_lower:
            resolved = all_schema_names_lower[new_schema.lower()]
            pairs.append((schema, old_cfg, resolved, schema_to_cfg[resolved]))
        else:
            missing_new.append((schema, old_cfg))

    return pairs, missing_new


def _query_one_schema(conn, schema: str) -> pd.DataFrame:
    """
    Query mnr_admin_area for AA1 features.
    Returns: feat_id, artificial, feat_type, name, standard_lang,
             country_code_char3, a1_admin_code
    """
    sql = f"""
        SELECT
            feat_id,
            artificial,
            feat_type,
            name,
            standard_lang,
            country_code_char3,
            a1_admin_code
        FROM {schema}.mnr_admin_area
        WHERE artificial IS FALSE
          AND feat_type = '1112'
        ;
    """
    return pd.read_sql_query(sql, conn)


def _analyse_pair(
    old_cfg: DBConfig,
    schema_old: str,
    new_cfg: DBConfig,
    schema_new: str,
    timestamp: str,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Run name + official_code comparison for one (old, new) schema pair.
    Returns (matched, name_error, missing_feature).
    """
    conn_old = _connect(old_cfg)
    try:
        df_old = _query_one_schema(conn_old, schema_old)
    finally:
        conn_old.close()

    conn_new = _connect(new_cfg)
    try:
        df_new = _query_one_schema(conn_new, schema_new)
    finally:
        conn_new.close()

    # Rename old columns
    df_old = df_old.rename(columns={
        "name":               "name_old",
        "standard_lang":      "standard_lang_new",
        "country_code_char3": "country_code",
        "a1_admin_code":      "OldOfficialCode",
    })

    # Keep only needed columns from new
    df_new = df_new[["feat_id", "name", "standard_lang", "a1_admin_code"]].rename(columns={
        "name":          "name_new",
        "standard_lang": "standard_lang_new_n",
        "a1_admin_code": "NewOfficialCode",
    })

    merged = df_old.merge(df_new, on="feat_id", how="left")

    # Use new standard_lang where available, fall back to old
    merged["standard_lang_new"] = merged["standard_lang_new_n"].combine_first(
        merged["standard_lang_new"]
    )
    merged.drop(columns=["standard_lang_new_n"], inplace=True)
    merged["_timestamp"] = int(timestamp)

    has_new    = merged["name_new"].notna()

    # Match: name AND official_code both identical
    name_match = merged["name_old"] == merged["name_new"]
    code_match = (
        merged["OldOfficialCode"].fillna("").astype(str) ==
        merged["NewOfficialCode"].fillna("").astype(str)
    )
    both_match = name_match & code_match

    matched         = merged[has_new &  both_match].copy()
    name_error      = merged[has_new & ~both_match].copy()
    missing_feature = merged[~has_new].copy()
    missing_feature["name_new"]       = missing_feature["name_new"].fillna("")
    missing_feature["NewOfficialCode"] = missing_feature["NewOfficialCode"].fillna("")

    def _order(df):
        for col in OUTPUT_COLS:
            if col not in df.columns:
                df[col] = ""
        return df[OUTPUT_COLS].reset_index(drop=True)

    return _order(matched), _order(name_error), _order(missing_feature)


def _analyse_old_only(
    old_cfg: DBConfig,
    schema_old: str,
    timestamp: str,
) -> pd.DataFrame:
    """
    New schema does not exist on any server.
    All features from old schema go to missing_feature.
    """
    conn = _connect(old_cfg)
    try:
        df_old = _query_one_schema(conn, schema_old)
    finally:
        conn.close()

    df_old = df_old.rename(columns={
        "name":               "name_old",
        "standard_lang":      "standard_lang_new",
        "country_code_char3": "country_code",
        "a1_admin_code":      "OldOfficialCode",
    })
    df_old["name_new"]        = ""
    df_old["NewOfficialCode"] = ""
    df_old["_timestamp"]      = int(timestamp)

    def _order(df):
        for col in OUTPUT_COLS:
            if col not in df.columns:
                df[col] = ""
        return df[OUTPUT_COLS].reset_index(drop=True)

    return _order(df_old)


def run_analysis(
    db_cfg: DBConfig,
    schema_old: str,
    schema_new: str,
    country_code: Optional[str] = None,
    progress_callback=None,
    db_cfg2: Optional[DBConfig] = None,
) -> AnalysisResult:
    """
    Main entry point.

    schema_old / schema_new are VERSION PREFIXES (e.g. _2026_06_004).
    db_cfg  = primary server   (caprod-cpp-pgmnr-001 -- EUR/CIS)
    db_cfg2 = secondary server (caprod-cpp-pgmnr-002 -- Americas/APAC/Africa)
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d")

    configs = [cfg for cfg in [db_cfg, db_cfg2] if cfg is not None]

    pairs, missing_new = discover_schema_pairs_multi(
        configs, schema_old, schema_new, country_code
    )

    if not pairs and not missing_new:
        raise ValueError(
            f"No schemas found on any server starting with '{schema_old}'. "
            f"Check the version prefix (e.g. _2026_06_004)."
        )

    total = len(pairs) + len(missing_new)

    all_matched = []
    all_errors  = []
    all_missing = []

    for i, (s_old, cfg_old, s_new, cfg_new) in enumerate(pairs):
        if progress_callback:
            progress_callback(i, total, s_old, s_new)
        m, e, miss = _analyse_pair(cfg_old, s_old, cfg_new, s_new, timestamp)
        all_matched.append(m)
        all_errors.append(e)
        all_missing.append(miss)

    base = len(pairs)
    for j, (s_old, cfg_old) in enumerate(missing_new):
        if progress_callback:
            progress_callback(base + j, total, s_old, "(new schema not found)")
        miss = _analyse_old_only(cfg_old, s_old, timestamp)
        all_missing.append(miss)

    def _concat(frames):
        non_empty = [f for f in frames if not f.empty]
        if not non_empty:
            return pd.DataFrame(columns=OUTPUT_COLS)
        combined = pd.concat(non_empty, ignore_index=True)
        # Some countries (e.g. CHN, VNM) have schemas on both servers.
        # Drop exact duplicate rows that result from querying the same data twice.
        return combined.drop_duplicates().reset_index(drop=True)

    return AnalysisResult(
        matched             = _concat(all_matched),
        name_error          = _concat(all_errors),
        missing_feature     = _concat(all_missing),
        timestamp           = timestamp,
        schemas_processed   = [(s_old, s_new) for s_old, _, s_new, _ in pairs],
        schemas_new_missing = [s_old for s_old, _ in missing_new],
    )


def to_excel_bytes(result: AnalysisResult) -> bytes:
    """Serialize result to Excel bytes with three sheets (matches FME output)."""
    import io
    from openpyxl.styles import Font, PatternFill, Alignment

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in [
            ("All_Matched_AA1",  result.matched),
            ("Name_Error",       result.name_error),
            ("Missing_feature",  result.missing_feature),
        ]:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]
            header_fill = PatternFill("solid", fgColor="1A56DB")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf.seek(0)
    return buf.read()
